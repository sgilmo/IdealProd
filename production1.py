#!/usr/bin/env python

"""Update Production Monitors."""

import os
import csv
import ftplib
import CommonFunc
from datetime import datetime
from datetime import date
from openpyxl import load_workbook


class ProdData(object):
    """Define Data Class."""

    def __init__(self):
        """Define Data Objects."""
        self.depts = []
        self.targets = []
        self.totals = []
        self.week = 0
        self.date = '1-1-99'
        self.dow = 0

    def add_dept(self, dept):
        """Set Dept."""
        self.depts.append(dept)

    def add_target(self, target):
        """Set Target."""
        self.targets.append(target)

    def add_total(self, total):
        """Set Total."""
        self.totals.append(total)


# Establish Some Globals
LIST_COMP = []
LIST_ASSY = []
LIST_PACK = []
LIST_SCREW = []
DAY = []
FILEDATA = {}


def step(ext, dirname, names):
    """Function for Directory Walk"""
    ext = ext.lower()
    for name in names:
        if name.lower().endswith(ext) and not name.startswith('~'):
            datename = datetime.strptime(name[:-5], '%m-%d-%y')
            FILEDATA[datename] = dirname + "\\" + name
    return


def getlisting(topdir):
    """Get Full Path File Listings of last 8 days"""
    # print 'File Name: ' + file_name[54:]
    keylist = []
    retlist = []
    exten = '.xlsm'
    for root, dirs, files in os.walk(topdir):
        step(exten, root, files)
    for key in sorted(FILEDATA.keys()):
        keylist.append(key)
    for item in keylist[-8:]:
        retlist.append(FILEDATA.get(item))
    return retlist


def buildfilelist(filelist):
    """Build List of Files to be displayed."""
    i = 0
    print(filelist)
    for item in filelist:
        DAY.append(ProdData())
        loaddata(item, DAY[i])
        i += 1
    return


def getdate(filename):
    """Strip date from file path\name."""
    temp = filename.split("\\")
    strdate = temp[-1][:-5]
    return strdate


def loaddata(file_name, day):
    """Get Data From Workbooks."""
    print("Opening Workbook for parsing", file_name)
    workbook = load_workbook(file_name, data_only=True)
    worksheet = workbook['Total summary']

    try:
        # Get date from file name
        date_sheet = datetime.strptime(getdate(file_name), '%m-%d-%y')
        day.date = date_sheet
        day.week = date_sheet.strftime("%W")
        day.dow = date_sheet.strftime("%w")
    except Exception as e:
        msg = 'Production Board Update Failed: ' + file_name + ': ' + str(e)
        CommonFunc.send_text('sgilmo', ['sgilmo'], msg)
    for row_num in range(4, 35):
        day.add_dept(worksheet.cell(row=row_num, column=1).value)
        day.add_target(worksheet.cell(row=row_num, column=3).value)
        day.add_total(worksheet.cell(row=row_num, column=16).value)
    return


def build_csv(file_name, data_list):
    """Build csv files"""
    with open(file_name, "w", newline='') as ftemp:
        writer = csv.writer(ftemp, delimiter=',')
        writer.writerow(data_list)
    return


def update_boards(filepath, ipaddr, area):
    """Update Scoreboards via FTP."""
    print("Transferring Board Data:  " + area)
    try:
        ftp_connect = ftplib.FTP(ipaddr, 'anonymous', 'anonymous')
        file_connect = open(filepath, "rb")
        ftp_connect.storbinary('STOR ' + area + '.csv', file_connect)
        file_connect.close()
        ftp_connect.quit()
    except Exception as e:
        msg = area + " Connection Failed: " + str(e)
        CommonFunc.send_text('sgilmo', ['sgilmo'], msg)
    return


# noinspection DuplicatedCode,DuplicatedCode
def fill_data_struct():
    """Fill Data Structures for each Production Area."""

    # Establish Board Destinations Tuples

    # Components
    board_comp = (17, 18, 19, 20, 21, 22, 23, 24, 25)
    # Assembly
    board_assy = (0, 1, 2, 4)
    # Packing Area
    board_pack = (3, 8, 9, 10, 11, 12, 13, 14, 15, 16)
    # Screw Department
    board_screw = (25, 26, 28, 29, 30)

    # Establish Week Number
    this_week = int(date.today().isocalendar()[1]) - 1  # -1

    target_sum = [0.00] * 40
    prod_sum = [0.00] * 40
    dept_name = ['Empty'] * 40
    today_dow = (datetime.today().weekday()) + 1  # +1
    # Load Data Structures
    for item in DAY:
        if int(item.week) < this_week or int(item.dow) == today_dow:
            continue
        for index in board_comp:
            dept_name[index] = item.depts[index]
            target_sum[index] = target_sum[index] + item.targets[index]
            prod_sum[index] = prod_sum[index] + item.totals[index]
        for index in board_assy:
            dept_name[index] = item.depts[index]
            target_sum[index] = target_sum[index] + item.targets[index]
            prod_sum[index] = prod_sum[index] + item.totals[index]
        for index in board_pack:
            dept_name[index] = item.depts[index]
            try:
                target_sum[index] = target_sum[index] + item.targets[index]
            except ValueError:
                print(item.week)
                print(item.depts)
                print(item.dow)
                print(index)
                print(target_sum[index])
                print(item.targets[index])
            prod_sum[index] = prod_sum[index] + item.totals[index]
        for index in board_screw:
            dept_name[index] = item.depts[index]
            target_sum[index] = target_sum[index] + item.targets[index]
            prod_sum[index] = prod_sum[index] + item.totals[index]

    # Build up Area Lists for Export to Screens
    for index in board_comp:
        LIST_COMP.append(dept_name[index])
        if target_sum[index] > 0:
            LIST_COMP.append(round((prod_sum[index] / target_sum[index]) * 100, 2))
        else:
            LIST_COMP.append(0.00)
    for index in board_assy:
        LIST_ASSY.append(dept_name[index])
        if target_sum[index] > 0:
            LIST_ASSY.append(round((prod_sum[index] / target_sum[index]) * 100, 2))
        else:
            LIST_ASSY.append(0.00)
    for index in board_pack:
        LIST_PACK.append(dept_name[index])
        if target_sum[index] > 0:
            LIST_PACK.append(round((prod_sum[index] / target_sum[index]) * 100, 2))
        else:
            LIST_PACK.append(0.00)
    for index in board_screw:
        LIST_SCREW.append(dept_name[index])
        if target_sum[index] > 0:
            LIST_SCREW.append(round((prod_sum[index] / target_sum[index]) * 100, 2))
        else:
            LIST_SCREW.append(0.00)
    return


def send_to_boards(path):
    """Send Data Files to Appropriate Board"""

    # Setup Dictionaries for Board Location and Data Lists
    area_ips = {'assy': '10.143.50.188', 'comp': '10.143.50.189',
                'pack': '10.143.50.190', 'screw': '10.143.50.191'}
    area_lists = {'assy': LIST_ASSY, 'comp': LIST_COMP,
                  'pack': LIST_PACK, 'screw': LIST_SCREW}

    # Build Data Files and Send to Boards
    for area, area_ip in area_ips.items():
        build_csv(path + area + ".csv", area_lists[area])
        update_boards(path + area + ".csv", area_ip, area)

    return


def main():
    """Main Function."""

    prodfilepath = '\\\\tn-san-fileserv\\Operator Output\\Daily Production Sheet\\' + str(datetime.now().year) + '\\'
    outputpath = 'c:\\temp\\'

    # Build File List of Appropriate Days
    buildfilelist(getlisting(prodfilepath))

    # Build Local Data Structures
    fill_data_struct()

    # Send Data Files to Appropriate Boards
    send_to_boards(outputpath)

    return


if __name__ == '__main__':
    main()
